# __Author__: Khoi Van 2021

import argparse
import json
import logging
import re
from itertools import zip_longest
from pathlib import Path, PurePath
from typing import Dict, List, Set, Union

import pandas as pd
from openpyxl import load_workbook
from rich.console import Console
from rich.logging import RichHandler
from rich.progress import Progress, BarColumn, SpinnerColumn, TimeElapsedColumn

# from upload import write_csv_to_google_sheet


console = Console()
# sys.setrecursionlimit(20000)

# Set logger using Rich: https://rich.readthedocs.io/en/latest/logging.html
logging.basicConfig(
    level="INFO",
    format="%(message)s",
    datefmt="[%X]",
    handlers=[RichHandler(rich_tracebacks=True)]
)
log = logging.getLogger("rich")


CURRENT_FILEPATH = Path(__file__).resolve().parent
DATA_FOLDER = CURRENT_FILEPATH / 'data'
# DATA_FOLDER.mkdir(exist_ok=True)
ORIGINAL_DATA_FOLDER = DATA_FOLDER / 'original'
BUILD_DATA_FOLDER = DATA_FOLDER / '_build'
BUILD_DATA_FOLDER.mkdir(exist_ok=True)

PRICEBOOK_FILE = ORIGINAL_DATA_FOLDER / 'Napoleon 2021-sanitized.xlsx'
NAPOLEON_CRUDE_DATA_FILE = BUILD_DATA_FOLDER / 'napoleon-crude-data.json'
NAPOLEON_DATABASE_FILE = BUILD_DATA_FOLDER / 'napoleon-database.json'

OPTIONAL_LOOKUP = {'mandatory': 'Required',
                   'optional': 'Optional'}
ADDITIONAL_OPTIONAL_LOOKUP = {'INC': 'Included',
                              'OPT': 'Optional',
                              'N/A': 'Not Available'}


def init_argparse() -> argparse.ArgumentParser:
    """Creating CLI helper"""
    parser = argparse.ArgumentParser(
        usage="python %(prog)s [OPTIONS]",
        description="Validate North Country Fire data for Napoleon dataset."
    )
    parser.add_argument('-d', '--debug',
                        help='Print more debug info.',
                        action="store_true")
    parser.add_argument('-r', '--reload-database',
                        help='Force reloading of database from pricebook.',
                        action="store_true")
    return parser


def extract_napoleon_data_from_catalog() -> Dict[str, Dict]:
    """Create a local database using the pricebook xlsx file

    Returns
    -------
    Dict[str, Dict]
        The database Dict object, consisting of three key, value pairs of data.
        The keys are: **series**, **variations**, and **products**
    """

    file = PRICEBOOK_FILE
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    catalog = {'series': {}, 'variations': {}, 'products': {}}
    series = {'content': []}
    # Using the values_only because you want to return the cells' values
    current_series = 0
    type = ''
    required_or_optional = ''
    variation_product_category = ''
    additional_options_baseSku = []

    for row in sheet.iter_rows(min_row=32,
                            #    max_row=sheet.max_rows,
                            #    values_only=True
                               ):
        row_data = [cell.value for cell in row if cell.value]

        # !DEBUG
        # if row[0].row >= 1170:
        #     breakpoint()

        # Pass on empty row
        if (not row_data
            # or re.match('FEATURES', row_data[0])
            or row_data[0].startswith('›')  # Rows in the Feature section normally startswith this symbol
            ):
            continue

        # Add a new series when there is a new 'features' cell
        # !WARNING: there is a typo in the pricebook on line 120 with 'FEATURES' is not at the beginning of the line
        if row_data and re.search('FEATURES', ''.join(str(cell) for cell in row_data)):
            current_series += 1
            catalog['series'][f'series-{current_series}'] = series
            series = {'content': []}
            continue

        # Signal the end of the pricebook:
        if row_data[0].lower().startswith('product returns'):
            break

        # Add current series vent type
        if row_data and re.search('venting',
                                  ''.join(str(cell) for cell in row_data),
                                  flags=re.IGNORECASE):
            # breakpoint()
            venting_type = re.search(r'top or rear|top & rear|top|rear', ''.join(str(cell) for cell in row_data), flags=re.IGNORECASE)
            if venting_type:
                series['venting'] = venting_type[0]

        # Identify the 'Step' as mark under each table to distinguish type of item, e.g. 'unit', 'variation', or 'product'
        step_number = re.search(r'^step\s*(\d+)', row_data[0], re.IGNORECASE)
        if step_number and int(step_number[1]) == 1:
            series['title'] = re.search(r'step 1\s*-\s*(.*)\(', row_data[0], re.IGNORECASE)[1].strip()
            type = 'unit'

        # Switch 'type' when encounter 'Step...'
        elif step_number and int(step_number[1]) > 1:
            type = 'variation'
            re_required_or_optional = re.compile(r'mandatory|optional', re.IGNORECASE)
            required_or_optional_result = re_required_or_optional.search(''.join(str(cell) for cell in row_data))
            if required_or_optional_result:
                matched_text = required_or_optional_result[0].lower()
                required_or_optional = OPTIONAL_LOOKUP.get(matched_text, '')
            product_category_match = re.search(r'^step\s*(\d+)(?:\W*)(.*)\(.*',
                                               ''.join(str(cell) for cell in row_data),
                                               re.IGNORECASE)
            if product_category_match:
                variation_product_category = product_category_match[2].strip()

        # Set type for 'additional_variations' for variations such as those in line 136-166 in Napoleon pricebook
        # for anomoly such as line 783 'Additional Vertical Series Option' instead of 'Vertical Series Additional Options'
        elif (len(row_data) == 1
              and not re.search(r'^step\s*(\d+)', row_data[0], re.IGNORECASE)    # For anomoly like line 738, containing 'Additional options' but startswith 'Step'
              and re.search(r'additional.*options', row_data[0], re.IGNORECASE | re.MULTILINE)):
            type = 'additional_variation'

        # Reset 'type' when reaching cells contains text: '... Design Options'
        elif len(row_data) == 1 and row_data[0].lower().endswith('design options'):
            type = ''

        # Set 'type' to 'product' when reaching cells having text ends with: '... components'
        elif len(row_data) == 1 and row_data[0].lower().endswith('components'):
            type = 'product'

        # Add 'unit' (fireplace, stove, insert)
        elif len(row_data) >= 1 and type == 'unit':
            if row_data[0].lower().startswith('product description'):
                series['baseSku'] = [baseSku.split('\n')[0] for baseSku in row_data[1:]]
                series['units'] = []
            else:
                name, *details = row_data
                info = []
                for line in details:
                    # console.log(f'Row #: {row[0].row}')
                    if (line.lower() != 'n/a'
                        and '\n' in line
                        and 'venting' not in line.lower()):
                        price, manufacturerSku = line.split('\n')
                        info.append({'price': price, 'manufacturerSku': manufacturerSku})
                    else:
                        info.append({})
                unit = {
                    'name': name,
                    'details': info,
                }
                series['units'].append(unit)

        # Add 'variations'
        elif len(row_data) > 1 and type == 'variation':
            # Add variations for each series
            if not series.get('variations'):
                series['variations'] = []

            # Get all of the SKU for units in this current series
            all_parent_sku_in_current_series = set([i['manufacturerSku']
                                                    for unit in series['units']
                                                    for i in unit['details']
                                                    if i])

            # For actual 'variation' row, not header row such as row containing 'Product description...'
            if (not row_data[0].lower().startswith('product description')
                and not re.search('FEATURES', row_data[0])):
                is_feature = False
                name, *details = row_data
                info = []

                # For each row of 'variation'
                for index, cell in enumerate(details):
                    if cell.lower() != 'n/a' and '\n' in cell:
                        price, manufacturerSku, *rest = cell.split('\n')

                        # If the price does not contain '$', this cell is not a real variations, so remove:
                        if not re.search(r'^\$', price):
                            is_feature = True
                            break

                        # Clean up pricebook input of baseSku with '\n' inside such as line 1008
                        # Fix for parentSku cannot be found when baseSku such as: 'GDI3N\nMillivolt Ingnition'
                        baseSku = series['baseSku'][index].split('\n')[0] if series['baseSku'] else None
                        parentSku = [sku for sku in all_parent_sku_in_current_series if sku.startswith(str(baseSku))] if baseSku else all_parent_sku_in_current_series

                        info.append({'price': price,
                                     'manufacturerSku': manufacturerSku,
                                     'base_sku': baseSku,
                                     'type': type,
                                    #  'parentSku': parentSku, #!Error: need to revisit later
                                     'requiredOrOptional': required_or_optional
                                     })

                        # Add variations directly into `catalog['variations']` for easy accessing:
                        variation_parents = list(zip_longest(parentSku,
                                                             [required_or_optional],
                                                             fillvalue=required_or_optional))
                        if not catalog['variations'].get(manufacturerSku):
                            catalog['variations'][manufacturerSku] = {
                                'name': name,
                                'catalog_product_category': variation_product_category,
                                'price': price,
                                'manufacturerSku': manufacturerSku,
                                'type': type,
                                'baseSku': [baseSku],
                                'variation_parents': variation_parents,
                            }
                        else:
                            catalog['variations'][manufacturerSku]['baseSku'].append(baseSku)
                            catalog['variations'][manufacturerSku]['variation_parents'].extend(variation_parents)

                # Append 'variations' to the current series
                if info and not is_feature:
                    variation = {
                        'name': name,
                        'catalog_product_category': variation_product_category,
                        'details': info,
                    }
                    series['variations'].append(variation)

        # Set variation_product_category for 'additional_variations' for variations such as those in line 136-166 in Napoleon pricebook
        # for anomoly such as line 783 'Additional Vertical Series Option' instead of 'Vertical Series Additional Options'
        elif (len(row_data) == 1
              and not re.search(r'^step\s*(\d+)', row_data[0], re.IGNORECASE)    # For anomoly like line 738, containing 'Additional options' but startswith 'Step'
              and type == 'additional_variation'):
            row_text = ''.join(str(cell) for cell in row_data)
            if row_text.isupper():
                variation_product_category = row_text.strip()

        # Add 'additional_variations'
        elif type == 'additional_variation':
            # console.log(f'Row #: {row[0].row}')
            # if row[0].row >= 149:
            #     breakpoint()

            allParentSku = {i['manufacturerSku']
                            for series in catalog['series'].values()
                            for unit in series['units']
                            for i in unit['details']
                            if i}
            if row_data[0].lower().startswith('product description'):
                _, _, _, *additional_options_baseSku = row_data
                # Clean up pricebook input of baseSku with '\n' inside such as line 1008
                additional_options_baseSku = [item.split('\n')[0]
                                              for item in additional_options_baseSku]
            elif len(row_data) >= 3:
                # Convert to string to fix problem with some partID (manufacturerSku) is interpreted as number instead of string, such line 1455
                fullname, manufacturerSku, price, *details = [str(cell) for cell in row_data]

                # For cases with multiple manufacturerSku inside a 'manufacturerSku' cell:
                # e.g: "Amber (MKBA), Black (MKBK), Blue (MKBB), Clear (MKBC), Topaz (MKBT)"
                re_multiple_sku_in_one_cell = re.compile(r"(\w*?)\s+\((.*?)\)", re.MULTILINE)

                matches = re_multiple_sku_in_one_cell.findall(manufacturerSku)
                # Split the matches if exist, for each match, add/modify a 'additional_variation' in `catalog['variations']`
                if matches:
                    for short_name, sku in matches:
                        catalog = add_additional_option_cell(details=details,
                                                             additional_options_baseSku=additional_options_baseSku,
                                                             allParentSku=allParentSku,
                                                             catalog=catalog,
                                                             manufacturerSku=sku,
                                                             extra_info={'name': f'{fullname}: {short_name}',
                                                                         'price': price,
                                                                         'type': type,
                                                                         'catalog_product_type': variation_product_category,
                                                                         }
                                                             )

                else:
                    catalog = add_additional_option_cell(details=details,
                                                         additional_options_baseSku=additional_options_baseSku,
                                                         allParentSku=allParentSku,
                                                         catalog=catalog,
                                                         manufacturerSku=manufacturerSku,
                                                        #  name=name, price=price, type=type
                                                         extra_info={'name': fullname,
                                                                     'price': price,
                                                                     'type': type,
                                                                     'catalog_product_type': variation_product_category,
                                                                     }
                                                         )

        # Add 'product'
        elif type == 'product':
            if (not row_data[0].lower().startswith('product description')
                and len(row_data) >= 3):
                # Convert to string to fix problem with some partID (manufacturerSku) is interpreted as number instead of string, such line 1455
                name, manufacturerSku, price, *_ = [str(cell) for cell in row_data]
                if not catalog['products'].get(manufacturerSku):
                    catalog['products'][manufacturerSku] = {
                        'name': name,
                        'price': price,
                        'manufacturerSku': manufacturerSku,
                        'type': type,
                    }
                else:
                    console.log(f'There is extra info for product {manufacturerSku}')

        # Add all necessary lines to the current series, `series['content']`  for debugging purpose
        if (row_data
            and not (len(row_data) == 1 and row_data[0].lower().startswith('bookmark'))
            and not re.search('FEATURES', row_data[0])):
            # console.print(row_data)
            series['content'].append(row_data)

    # Add to catalog the last item (because there is no 'features' cell at the end)
    current_series += 1
    catalog['series'][f'series-{current_series}'] = series

    # Update all Additional Options to include more series
    # since not all series has been added
    return update_additional_options(catalog)


def add_additional_option_cell(details: List[str],
                               additional_options_baseSku: List[str],
                               allParentSku: Set[str],
                               catalog: Dict[str, Dict],
                               manufacturerSku: str,
                            #    name: str,
                            #    price: Union[int, str],
                            #    type: str,
                               extra_info: Dict[str, str],
                               ) -> Dict[str, Dict]:
    """Add item that marked as 'additional_option' into local database

    The 'additional_option' are just 'variation'

    Parameters
    ----------
    details : List[str]
        Generally 'Optional', 'Included', or 'Not Available'
    additional_options_baseSku : List[str]
        List of baseSku provided by the header of the table
    allParentSku : Set[str]
        Set of unique parentSku
    catalog : Dict[str, Dict]
        The local database
    manufacturerSku : str
        manufacturerSku of this item
    name : str
        name of the item
    price : Union[int, str]
        price of the item
    type : str
        the type of the item

    Returns
    -------
    Dict[str, Dict]
        the catalog (database) Dict object
    """
    for index, cell in enumerate(details):
        baseSku = additional_options_baseSku[index]
        parentSku = {sku for sku in allParentSku if sku.startswith(baseSku)}

        # For those baseSku that the series (full parentSku) does not exist:
        # save the baseSku as a placeholder
        # and add 'need update' as the second element of a tuple to signal for update later
        if not parentSku:
            parentSku.add((baseSku, 'need update'))

        additional_option_requirement = ADDITIONAL_OPTIONAL_LOOKUP[cell]

        # Add variations for series as for better congregation:
        variation_parents = list(zip_longest(parentSku,
                                             [additional_option_requirement],
                                             fillvalue=additional_option_requirement))
        if not catalog['variations'].get(manufacturerSku):
            catalog['variations'][manufacturerSku] = {
                'manufacturerSku': manufacturerSku,
                # 'name': name,
                # 'price': price,
                # 'type': type,
                **extra_info,
                'baseSku': [baseSku],
                'variation_parents': variation_parents,
            }
        else:
            catalog['variations'][manufacturerSku]['baseSku'].append(baseSku)
            catalog['variations'][manufacturerSku]['variation_parents'].extend(variation_parents)

    return catalog


def update_additional_options(database: Dict[str, Dict]) -> Dict[str, Dict]:
    """Update all Additional Options to include more series
    since not all series has been added

    This operation has to run after all series have been added, aka
    after the local database (`catalog` Dict object)
    has been populated from the pricebook

    Parameters
    ----------
    database : Dict[str, Dict]
        The local database that needs updated

    Returns
    -------
    Dict[str, Dict]
        The database with updated parentSku
    """
    all_unit_sku = get_all_unit_sku(database=database)
    # console.log(all_unit_sku)
    for variation_info in database['variations'].values():
        # !DEBUG
        # if variation_info['manufacturerSku'] == "BKGDIX3":
        #     breakpoint()

        # Filter out any just (baseSku, condition) in the 'variation_parents' list of tuple
        # Any parent_sku need updated is in a form of `(('parentSku', 'need updated'), condition)`
        # example: `(('GDIZC', 'need update'), 'Not Available'), (('GDIZC', 'need update'), 'Not Available')`
        need_updated_parent_sku = [(sku, condition)
                                   for sku, condition in variation_info['variation_parents']
                                   if 'need update' in sku
                                   ]
        new_variation_parents = [item for item in variation_info['variation_parents']
                                 if item not in need_updated_parent_sku]
        for baseSku, condition in need_updated_parent_sku:
            parentSku = [sku for sku in all_unit_sku if sku.startswith(baseSku)]
            new_variation_parents.extend(list(zip_longest(parentSku, [condition], fillvalue=condition)))
        variation_info['variation_parents'] = new_variation_parents
    return database


def get_all_unit_sku(database: Dict[str, Dict]) -> Set[str]:
    """Return a set of ALL 'units' manufacturerSku in the current database/catalog

    Parameters
    ----------
    database : Dict[str, Dict]
        Current database

    Returns
    -------
    Set[str]
        A unique set of string of all units' manufacturerSku
    """
    return {
        i['manufacturerSku']
        for series in database['series'].values()
        for unit in series['units']
        for i in unit['details']
        if i
    }


def validate_ncf(file_to_validate: PurePath,
                 database: Dict[str, Dict],
                 target_file:  PurePath = None) -> None:
    """Main function to validate the North Country Fire template file

    Parameters
    ----------
    file_to_validate : PurePath
        file path to the template file
    database : Dict[str, Dict]
        the current database/catalog to look up info
    target_file : PurePath, optional
        file path for the desired validated output, by default None
    """
    # Set default validated output file if not set
    if not target_file:
        target_file = VALIDATED_NCF_FILE

    # # !Using excel have consequences of Excel interpreting data in the different way,
    # # such as '2200-1' is thought (by Excel, or google sheet as date format)
    # excel_file = file_to_validate
    # ncf_data = pd.read_excel(excel_file, dtype=object)

    # !Use csv input ncf template file instead
    ncf_data = pd.read_csv(file_to_validate, dtype=object)

    # Fill SKU (same as manufacturerSku)
    ncf_data['c__sku'] = ncf_data['manufacturerSKU']

    # Validate 'c__unitTrueOrFalse'
    ncf_data['c__unitTrueOrFalse'] = ncf_data.apply(is_unit,
                                                    axis=1,     # Apply for column
                                                    args=(database,))

    # Validate baseSku
    ncf_data['c__baseSku'] = ncf_data.apply(check_base_sku,
                                            axis=1,
                                            args=(database,))

    # Validate 'parentSku'
    ncf_data['c__parentSku'] = ncf_data.apply(check_parent_sku,
                                              axis=1,
                                              args=(database,))

    # Validate 'c__isSharedVariationProduct', run AFTER `check_parent_sku`
    ncf_data['c__isSharedVariationProduct'] = ncf_data.apply(is_shared_variation_product,
                                                             axis=1)

    # Validate 'c__isStepVariationProduct', run AFTER `check_parent_sku`
    ncf_data['c__isStepVariationProduct'] = ncf_data.apply(is_step_variation_product,
                                                           axis=1)

    # Validate 'c__isStepVariationProduct', run AFTER `is_step_variation_product`
    ncf_data['c__requiredOrOptionalVariation'] = ncf_data.apply(required_or_optional_variation,
                                                                axis=1,
                                                                args=(database,))

    # Save the validated result
    file_extension = target_file.suffix
    if file_extension == '.csv':
        # Save to csv file for easy checking
        ncf_data.to_csv(target_file, index=False)
    else:
        ncf_data.to_excel(target_file, index=False)


def is_unit(row: pd.Series, database: Dict[str, Dict]) -> str:
    """Check manufacturerSku to see if the item is a 'unit'

    'Unit': fireplace, stove, or inserts

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`
    database : Dict[str, Dict]
        the local database/catalog

    Returns
    -------
    str
        'TRUE' if unit,
        'Not found' if item does not exist in database,
        '' (blank) if found but not a 'unit'
    """
    # To tolerate input typo (lower case) in ncf that causes issue: e.g,  in ncf file, line 748, 603, 543
    # Cannot fix with a simple `str.upper()` due to there is SKU such as 'S20i' and 'S25i'
    manufacturerSku = row['manufacturerSKU']
    if not re.search(r'[A-Z]', manufacturerSku):
        manufacturerSku = row['manufacturerSKU'].upper()
    result = ''
    variation = database['variations'].get(manufacturerSku)
    product = database['products'].get(manufacturerSku)
    if not variation and not product:
        for series in database['series'].values():
            if manufacturerSku in (i['manufacturerSku']
                                   for unit in series['units']
                                   for i in unit['details']
                                   if i):
                result = 'TRUE'
                break
        else:
            result = 'Not found'
    return result


def check_base_sku(row: pd.Series, database: Dict[str, Dict]) -> str:
    """Return the baseSku string for 'unit'

    Generally, the baseSku = first letters of manufacturerSku + first digits
    There are exceptions such as 'S20i', 'S25i', 'BHD4-Glass', 'BHD4-Cradle'

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`
    database : Dict[str, Dict]
        the local database/catalog

    Returns
    -------
    str
        The baseSku if 'unit',
        'Not found' if does not exist in database,
        '' (blank) if found but not a 'unit'
    """
    # To tolerate input typo (lower case) in ncf that causes issue: e.g,  in ncf file, line 748, 603, 543
    # Cannot fix with a simple `str.upper()` due to there is SKU such as 'S20i' and 'S25i'
    manufacturerSku = row['manufacturerSKU']
    if not re.search(r'[A-Z]', manufacturerSku):
        manufacturerSku = row['manufacturerSKU'].upper()
    baseSku = ''
    variation = database['variations'].get(manufacturerSku)
    product = database['products'].get(manufacturerSku)
    if not variation and not product:
        for series in database['series'].values():
            if manufacturerSku in (i['manufacturerSku']
                                   for unit in series['units']
                                   for i in unit['details']
                                   if i):
                # Most of the series have 'baseSku'
                if series['baseSku']:
                    # baseSku = ','.join(sku for sku in series['baseSku'] if manufacturerSku.startswith(sku))
                    baseSku = next((sku for sku in series['baseSku']
                                    if manufacturerSku.startswith(sku)),
                                   None)

                    # For cases such as pricebook lines 614-615. 'manufacturerSku': 'BHD4STFCN' <--> 'baseSku': 'BHD4-Cradle'
                    if not baseSku:
                        unit_details = next((unit
                                             for unit in series['units']
                                             for item in unit['details']
                                             if item and manufacturerSku == item['manufacturerSku']),
                                            None)
                        unit_index = series['units'].index(unit_details)
                        baseSku = series['baseSku'][unit_index]

                    # Special cases such as for ncf lines: 112, 113, 262, 372, 663, 821, 836, 837, 853, 637-40 and 727
                    # Example: 'BHD4-Glass', 'BHD4-Cradle', 'NEFB33H', 'NEFB40H',
                    # 'NEFBD50HE', 'NEFB36H-BS', 'GDIZC', 'GDI3N', 'GDIG3N',
                    # 'GDIX3N', 'GDIX4N', 'GD82NT-PA', 'GSS36CF', 'S20i',
                    # 'NEFP33-0214W', 'NEFB50H-3SV', 'NEFB60H-3SV',
                    # !IMPORTANT: leave these baseSku the way they are: 'BHD4-Glass', 'BHD4-Cradle', 'S20i'
                    if not baseSku[-1].islower():
                        baseSku = re.search(r'^[A-Z]*\d*', baseSku)[0]

                # For some rare case without 'baseSku', e.g. pricebook lines 1818-1819, 1841-1842
                else:
                    # console.log(manufacturerSku)
                    # See here for info on the regex: https://regex101.com/r/E9id2S/1/
                    baseSku = re.search(r'^[A-Z]*\d*', manufacturerSku)[0]

                break
        else:
            baseSku = 'Not found'
    return baseSku


def check_parent_sku(row: pd.Series, database: Dict[str, Dict]) -> str:
    """Return full parentSku for each item

    If variation = full unit manufacturerSku to which the variation product belongs
        i) If >1 unit, separate values by commas

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`
    database : Dict[str, Dict]
        the local database/catalog

    Returns
    -------
    str
        The baseSku if 'unit',
        '' (blank) if not 'variation' (e.g. 'unit', 'product')
    """
    # To tolerate input typo (lower case) in ncf that causes issue: e.g,  in ncf file, line 748, 603, 543
    # Cannot fix with a simple `str.upper()` due to there is SKU such as 'S20i' and 'S25i'
    manufacturerSku = row['manufacturerSKU']
    if not re.search(r'[A-Z]', manufacturerSku):
        manufacturerSku = row['manufacturerSKU'].upper()
    variation = database['variations'].get(manufacturerSku)
    parent_sku = ''
    if variation:
        parent_sku = ','.join(sorted({sku for sku, _ in variation['variation_parents']}))
    return parent_sku


def is_shared_variation_product(row: pd.Series) -> str:
    """Return if an item (not 'unit') have more than one parents

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`

    Returns
    -------
    str
        'TRUE' if number of parentSku > 1; '' (blank) otherwise
    """
    parent_sku = row['c__parentSku']
    more_than_one_parent = parent_sku.count(',') >= 1   # Only need one comma for 2 parents
    return 'TRUE' if more_than_one_parent else ''


def is_step_variation_product(row: pd.Series) -> str:
    """Return if an item (not 'unit') have at least one parent

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`

    Returns
    -------
    str
        'TRUE' if number of parentSku >= 1; '' (blank) otherwise
    """
    return 'TRUE' if row['c__parentSku'] else ''


def required_or_optional_variation(row: pd.Series, database: Dict[str, Dict]) -> str:
    """Return the requirement condition of the 'variation'

    a) If isStepVariationProduct = TRUE:
        i) If variation is required to complete unit(s) = 'Required'
        ii) If variation is optional to complete unit(s) = 'Optional'
        iii) If variation is required to complete some units and optional for other units
            (or even the same units, marked differently in the catalog)
            = 'Required/Optional'
        iv) If variation is included with some units and optional for others
             (or even the same units, marked differently in the catalog)
            = 'Included/Optional'
    b) If isStepVariationProduct is FALSE (blank) = blank

    Parameters
    ----------
    row : pd.Series
        pandas Series, passed in by the `pd.DataFrame.apply()`
    database : Dict[str, Dict]
        the local database/catalog

    Returns
    -------
    str
        'Required', 'Optional', 'Required/Optional', 'Included/Optional'
    """
    if row['c__isStepVariationProduct'] == 'TRUE':
        # To tolerate input typo (lower case) in ncf that causes issue: e.g,  in ncf file, line 748, 603, 543
        # Cannot fix with a simple `str.upper()` due to there is SKU such as 'S20i' and 'S25i'
        manufacturerSku = row['manufacturerSKU']
        if not re.search(r'[A-Z]', manufacturerSku):
            manufacturerSku = row['manufacturerSKU'].upper()

        variation = database['variations'].get(manufacturerSku)
        if variation:
            # Remove 'Not Available'
            requirements = {requirement
                            for _, requirement in variation['variation_parents']
                            if requirement != 'Not Available'
                            }
            value = sorted(requirements) if 'Included' in requirements else sorted(requirements, reverse=True)
            return '/'.join(value)
    return ''


if __name__ == "__main__":
    parser = init_argparse()
    debug = parser.parse_args().debug
    reload_db = parser.parse_args().reload_database

    database = {}

    progress = Progress(SpinnerColumn(),
                        "[magenta]{task.description}",
                        BarColumn(),
                        TimeElapsedColumn(),
                        console=console,
                        transient=True)

    if debug:
        # Reload the database using the pricebook
        if reload_db:
            log.info('[bold red blink]Regenerating database from pricebook. Please wait![/]', extra={"markup": True})
            database = extract_napoleon_data_from_catalog()

            # Save the new database into json file
            with open(NAPOLEON_CRUDE_DATA_FILE, 'w') as fp:
                json.dump(database, fp,  indent=2)

        # Otherwise, read from the json file
        else:
            log.info(f'Loading database from {NAPOLEON_CRUDE_DATA_FILE}.')
            with open(NAPOLEON_CRUDE_DATA_FILE, 'r') as fin:
                database = json.load(fin)
    else:
        with progress:
            # Reload the database using the pricebook
            if reload_db:
                log.info('[bold red blink]Regenerating database from pricebook. Please wait![/]', extra={"markup": True})
                task1 = progress.add_task('Creating database...', start=True)
                database = extract_napoleon_data_from_catalog()

                # Save the new database into json file
                with open(NAPOLEON_CRUDE_DATA_FILE, 'w') as fp:
                    json.dump(database, fp,  indent=2)

            # Otherwise, read from the json file
            else:
                log.info(f'Loading database from {NAPOLEON_CRUDE_DATA_FILE}.')
                task1 = progress.add_task('Loading database...', start=True)
                with open(NAPOLEON_CRUDE_DATA_FILE, 'r') as fin:
                    database = json.load(fin)

            progress.update(task1, completed=100)


    # Printing info about the database/catalog
    log.info('This database/catalog contains:')
    log.info(f"Number of series: {len(database['series'])}")
    log.info(f"Number of variations: {len(database['variations'])}")
    log.info(f"Number of products: {len(database['products'])}")

    # with console.status("[bold green]Validating NCF file...") as status:
    #     # Validate NCF file
    #     validate_ncf(file_to_validate=NCF_CSV_FILE,
    #                 database=database,
    #                 target_file=VALIDATED_NCF_CSV_FILE,
    #                 #  target_file=VALIDATED_NCF_FILE,
    #                 )
    #     console.log('NCF file is populated and validated!')


    # with console.status("[bold green]Uploading validated file to Google sheet...") as status:
    #     # Upload validated file to google sheet
    #     sheet_url = write_csv_to_google_sheet(VALIDATED_NCF_CSV_FILE)
    #     console.log('Validated file upload to googlesheet!')
    #     console.log(f'Sheet URL: {sheet_url}')

    # Print CLI helper if the code was not called with any argument
    if not (debug or reload_db):
        console.print('\n\nCLI info:', style='bold red')
        parser.print_help()